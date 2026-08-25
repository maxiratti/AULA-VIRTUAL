from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Prefetch, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from django.utils import timezone
from django.urls import reverse

from apps.actividades.models import Actividad, Entrega
from apps.contenidos.models import (
    Clase,
    Modulo,
    ProgresoClase,
)

from apps.inscripciones.models import Inscripcion
from apps.notificaciones.services import (
    notificar_aviso_curso,
    notificar_curso_finalizado,
    notificar_mensaje_curso,
)
from apps.roles.utils import tiene_rol_en_institucion

from .forms import CursoForm
from .models import (
    AvisoCurso,
    Certificado,
    ConversacionCurso,
    Curso,
    MensajeCurso,
)


def instituciones_coordinadas(usuario):
    if usuario.is_superuser:
        return None

    return (
        usuario.membresias
        .filter(
            activa=True,
            institucion__activa=True,
            roles__name="Coordinador",
        )
        .values_list(
            "institucion_id",
            flat=True,
        )
        .distinct()
    )


@login_required
def lista_cursos(request):
    if request.user.is_superuser:
        cursos = Curso.objects.all()

    else:
        instituciones_ids = instituciones_coordinadas(
            request.user
        )

        if not instituciones_ids:
            raise PermissionDenied

        cursos = Curso.objects.filter(
            institucion_id__in=instituciones_ids
        )

    cursos = list(
        cursos
        .select_related("institucion")
        .prefetch_related("docentes")
        .order_by(
            "institucion__nombre",
            "nombre",
        )
    )

    for curso in cursos:
        curso.mensajes_pendientes = (
            MensajeCurso.objects
            .filter(
                conversacion__curso=curso,
                autor__in=Inscripcion.objects
                .filter(curso=curso)
                .values("alumno"),
                leido_por_equipo=False,
            )
            .count()
        )

    return render(
        request,
        "cursos/lista.html",
        {
            "cursos": cursos,
        },
    )


@login_required
def nuevo_curso(request):
    if request.user.is_superuser:
        instituciones_permitidas = None

    else:
        instituciones_permitidas = (
            request.user.membresias
            .filter(
                activa=True,
                institucion__activa=True,
                roles__name="Coordinador",
            )
            .values_list(
                "institucion_id",
                flat=True,
            )
            .distinct()
        )

        if not instituciones_permitidas:
            raise PermissionDenied

    if request.method == "POST":
        form = CursoForm(request.POST)

        if instituciones_permitidas is not None:
            form.fields["institucion"].queryset = (
                form.fields["institucion"]
                .queryset
                .filter(
                    pk__in=instituciones_permitidas
                )
            )

        if form.is_valid():
            curso = form.save(commit=False)

            if (
                not request.user.is_superuser
                and not tiene_rol_en_institucion(
                    request.user,
                    "Coordinador",
                    curso.institucion,
                )
            ):
                raise PermissionDenied

            curso.save()
            form.save_m2m()

            messages.success(
                request,
                "Curso creado correctamente.",
            )

            return redirect("lista_cursos")

    else:
        form = CursoForm()

        if instituciones_permitidas is not None:
            form.fields["institucion"].queryset = (
                form.fields["institucion"]
                .queryset
                .filter(
                    pk__in=instituciones_permitidas
                )
            )

    return render(
        request,
        "cursos/form.html",
        {
            "form": form,
            "titulo": "Nuevo curso",
            "subtitulo": (
                "Creá un nuevo curso y asigná "
                "sus docentes."
            ),
        },
    )


@login_required
def editar_curso(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    if not request.user.is_superuser:
        if not tiene_rol_en_institucion(
            request.user,
            "Coordinador",
            curso.institucion,
        ):
            raise PermissionDenied

    if request.method == "POST":
        form = CursoForm(
            request.POST,
            instance=curso,
        )

        if not request.user.is_superuser:
            instituciones_permitidas = (
                request.user.membresias
                .filter(
                    activa=True,
                    institucion__activa=True,
                    roles__name="Coordinador",
                )
                .values_list(
                    "institucion_id",
                    flat=True,
                )
                .distinct()
            )

            form.fields["institucion"].queryset = (
                form.fields["institucion"]
                .queryset
                .filter(
                    pk__in=instituciones_permitidas
                )
            )

        if form.is_valid():
            curso_editado = form.save(commit=False)

            if (
                not request.user.is_superuser
                and not tiene_rol_en_institucion(
                    request.user,
                    "Coordinador",
                    curso_editado.institucion,
                )
            ):
                raise PermissionDenied

            curso_editado.save()
            form.save_m2m()

            messages.success(
                request,
                "Curso actualizado correctamente.",
            )

            return redirect("lista_cursos")

    else:
        form = CursoForm(
            instance=curso,
        )

        if not request.user.is_superuser:
            instituciones_permitidas = (
                request.user.membresias
                .filter(
                    activa=True,
                    institucion__activa=True,
                    roles__name="Coordinador",
                )
                .values_list(
                    "institucion_id",
                    flat=True,
                )
                .distinct()
            )

            form.fields["institucion"].queryset = (
                form.fields["institucion"]
                .queryset
                .filter(
                    pk__in=instituciones_permitidas
                )
            )

    return render(
        request,
        "cursos/form.html",
        {
            "form": form,
            "titulo": "Editar curso",
            "subtitulo": (
                "Actualizá los datos del curso."
            ),
        },
    )


def puede_gestionar_avisos(usuario, curso):
    return (
        usuario.is_superuser
        or curso.docentes.filter(pk=usuario.pk).exists()
        or tiene_rol_en_institucion(
            usuario,
            "Coordinador",
            curso.institucion,
        )
        or tiene_rol_en_institucion(
            usuario,
            "Administrador institucional",
            curso.institucion,
        )
    )


@login_required
def avisos_curso(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    if not puede_gestionar_avisos(request.user, curso):
        raise PermissionDenied

    if request.method == "POST":
        if curso.estado == Curso.ESTADO_FINALIZADO:
            messages.warning(
                request,
                "El curso está finalizado. Los avisos quedan en modo consulta.",
            )
            return redirect("avisos_curso", pk=curso.pk)

        titulo = request.POST.get("titulo", "").strip()
        mensaje = request.POST.get("mensaje", "").strip()
        visible = request.POST.get("visible") == "1"

        if not titulo or not mensaje:
            messages.error(
                request,
                "Completá el título y el mensaje del aviso.",
            )
        else:
            aviso = AvisoCurso.objects.create(
                curso=curso,
                autor=request.user,
                titulo=titulo,
                mensaje=mensaje,
                visible=visible,
            )

            if aviso.visible:
                notificar_aviso_curso(aviso)

            messages.success(
                request,
                (
                    "Aviso publicado correctamente."
                    if aviso.visible
                    else "Aviso guardado como oculto."
                ),
            )

            return redirect("avisos_curso", pk=curso.pk)

    avisos = (
        AvisoCurso.objects
        .filter(curso=curso)
        .select_related("autor")
    )

    return render(
        request,
        "cursos/avisos.html",
        {
            "curso": curso,
            "avisos": avisos,
        },
    )


@login_required
def editar_aviso_curso(request, curso_pk, aviso_pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_pk,
    )
    aviso = get_object_or_404(
        AvisoCurso,
        pk=aviso_pk,
        curso=curso,
    )

    if not puede_gestionar_avisos(request.user, curso):
        raise PermissionDenied

    if curso.estado == Curso.ESTADO_FINALIZADO:
        messages.warning(
            request,
            "El curso está finalizado. Los avisos quedan en modo consulta.",
        )
        return redirect("avisos_curso", pk=curso.pk)

    if request.method != "POST":
        return redirect("avisos_curso", pk=curso.pk)

    titulo = request.POST.get("titulo", "").strip()
    mensaje = request.POST.get("mensaje", "").strip()
    visible = request.POST.get("visible") == "1"

    if not titulo or not mensaje:
        messages.error(
            request,
            "Completá el título y el mensaje del aviso.",
        )
        return redirect("avisos_curso", pk=curso.pk)

    era_visible = aviso.visible

    aviso.titulo = titulo
    aviso.mensaje = mensaje
    aviso.visible = visible
    aviso.save(
        update_fields=[
            "titulo",
            "mensaje",
            "visible",
            "fecha_actualizacion",
        ]
    )

    if visible and not era_visible:
        notificar_aviso_curso(aviso)

    messages.success(
        request,
        "Aviso actualizado correctamente.",
    )

    return redirect("avisos_curso", pk=curso.pk)


@login_required
def eliminar_aviso_curso(request, curso_pk, aviso_pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_pk,
    )
    aviso = get_object_or_404(
        AvisoCurso,
        pk=aviso_pk,
        curso=curso,
    )

    if not puede_gestionar_avisos(request.user, curso):
        raise PermissionDenied

    if curso.estado == Curso.ESTADO_FINALIZADO:
        messages.warning(
            request,
            "El curso está finalizado. Los avisos quedan en modo consulta.",
        )
        return redirect("avisos_curso", pk=curso.pk)

    if request.method == "POST":
        aviso.delete()
        messages.success(
            request,
            "Aviso eliminado correctamente.",
        )

    return redirect("avisos_curso", pk=curso.pk)


def puede_gestionar_mensajeria(usuario, curso):
    return (
        usuario.is_superuser
        or curso.docentes.filter(pk=usuario.pk).exists()
        or tiene_rol_en_institucion(
            usuario,
            "Coordinador",
            curso.institucion,
        )
        or tiene_rol_en_institucion(
            usuario,
            "Administrador institucional",
            curso.institucion,
        )
    )


@login_required
def mensajeria_alumno(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    inscripcion = (
        Inscripcion.objects
        .filter(
            curso=curso,
            alumno=request.user,
        )
        .first()
    )

    if (
        not inscripcion
        or not tiene_rol_en_institucion(
            request.user,
            "Alumno",
            curso.institucion,
        )
    ):
        raise PermissionDenied

    conversacion, _ = ConversacionCurso.objects.get_or_create(
        curso=curso,
        alumno=request.user,
    )

    if request.method == "POST":
        if curso.estado == Curso.ESTADO_FINALIZADO:
            messages.warning(
                request,
                "El curso está finalizado. La conversación queda en modo consulta.",
            )
            return redirect("mensajeria_alumno", pk=curso.pk)

        texto = request.POST.get("mensaje", "").strip()

        if not texto:
            messages.error(
                request,
                "Escribí un mensaje antes de enviarlo.",
            )
        else:
            mensaje = MensajeCurso.objects.create(
                conversacion=conversacion,
                autor=request.user,
                mensaje=texto,
                leido_por_alumno=True,
                leido_por_equipo=False,
            )
            conversacion.save()
            notificar_mensaje_curso(mensaje)

            messages.success(
                request,
                "Consulta enviada correctamente.",
            )
            return redirect("mensajeria_alumno", pk=curso.pk)

    conversacion.mensajes.filter(
        leido_por_alumno=False,
    ).exclude(
        autor=request.user,
    ).update(
        leido_por_alumno=True,
    )

    mensajes_conversacion = (
        conversacion.mensajes
        .select_related("autor")
        .all()
    )

    return render(
        request,
        "cursos/mensajeria_alumno.html",
        {
            "curso": curso,
            "inscripcion": inscripcion,
            "conversacion": conversacion,
            "mensajes_conversacion": mensajes_conversacion,
        },
    )


@login_required
def mensajeria_docente(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    if not puede_gestionar_mensajeria(request.user, curso):
        raise PermissionDenied

    conversaciones = list(
        ConversacionCurso.objects
        .filter(curso=curso)
        .select_related("alumno")
        .prefetch_related("mensajes")
        .order_by("-fecha_actualizacion")
    )

    for conversacion in conversaciones:
        conversacion.pendientes = (
            conversacion.mensajes
            .filter(
                autor=conversacion.alumno,
                leido_por_equipo=False,
            )
            .count()
        )

    return render(
        request,
        "cursos/mensajeria_docente.html",
        {
            "curso": curso,
            "conversaciones": conversaciones,
        },
    )


@login_required
def mensajeria_docente_conversacion(
    request,
    curso_pk,
    alumno_pk,
):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_pk,
    )

    if not puede_gestionar_mensajeria(request.user, curso):
        raise PermissionDenied

    inscripcion = get_object_or_404(
        Inscripcion.objects.select_related("alumno"),
        curso=curso,
        alumno_id=alumno_pk,
    )

    conversacion, _ = ConversacionCurso.objects.get_or_create(
        curso=curso,
        alumno=inscripcion.alumno,
    )

    if request.method == "POST":
        if curso.estado == Curso.ESTADO_FINALIZADO:
            messages.warning(
                request,
                "El curso está finalizado. La conversación queda en modo consulta.",
            )
            return redirect(
                "mensajeria_docente_conversacion",
                curso_pk=curso.pk,
                alumno_pk=inscripcion.alumno.pk,
            )

        texto = request.POST.get("mensaje", "").strip()

        if not texto:
            messages.error(
                request,
                "Escribí una respuesta antes de enviarla.",
            )
        else:
            mensaje = MensajeCurso.objects.create(
                conversacion=conversacion,
                autor=request.user,
                mensaje=texto,
                leido_por_alumno=False,
                leido_por_equipo=True,
            )
            conversacion.save()
            notificar_mensaje_curso(mensaje)

            messages.success(
                request,
                "Respuesta enviada correctamente.",
            )
            return redirect(
                "mensajeria_docente_conversacion",
                curso_pk=curso.pk,
                alumno_pk=inscripcion.alumno.pk,
            )

    conversacion.mensajes.filter(
        autor=inscripcion.alumno,
        leido_por_equipo=False,
    ).update(
        leido_por_equipo=True,
    )

    mensajes_conversacion = (
        conversacion.mensajes
        .select_related("autor")
        .all()
    )

    return render(
        request,
        "cursos/mensajeria_conversacion.html",
        {
            "curso": curso,
            "inscripcion": inscripcion,
            "conversacion": conversacion,
            "mensajes_conversacion": mensajes_conversacion,
        },
    )


def verificar_certificado(request, codigo):
    certificado = get_object_or_404(
        Certificado.objects.select_related(
            "curso",
            "curso__institucion",
            "alumno",
        ),
        codigo=codigo,
    )

    alumno = certificado.alumno
    nombre_alumno = (
        alumno.get_full_name().strip()
        or alumno.username
    )

    fecha_final = (
        Inscripcion.objects
        .filter(
            curso=certificado.curso,
            alumno=alumno,
        )
        .values_list(
            "fecha_finalizacion",
            flat=True,
        )
        .first()
    )

    return render(
        request,
        "cursos/verificar_certificado.html",
        {
            "certificado": certificado,
            "nombre_alumno": nombre_alumno,
            "fecha_final": fecha_final,
        },
    )


@login_required
def certificado_alumno(request, curso_pk, alumno_pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_pk,
    )

    inscripcion = get_object_or_404(
        Inscripcion.objects.select_related(
            "alumno",
            "curso",
            "curso__institucion",
        ),
        curso=curso,
        alumno_id=alumno_pk,
    )

    es_el_alumno = request.user.pk == inscripcion.alumno_id
    puede_gestionar = (
        request.user.is_superuser
        or curso.docentes.filter(pk=request.user.pk).exists()
        or tiene_rol_en_institucion(
            request.user, "Coordinador", curso.institucion
        )
        or tiene_rol_en_institucion(
            request.user,
            "Administrador institucional",
            curso.institucion,
        )
    )

    if not es_el_alumno and not puede_gestionar:
        raise PermissionDenied

    if (
        curso.estado != Curso.ESTADO_FINALIZADO
        or inscripcion.estado != Inscripcion.ESTADO_APROBADO
    ):
        raise PermissionDenied

    alumno = inscripcion.alumno
    nombre_alumno = alumno.get_full_name().strip() or alumno.username
    fecha_final = inscripcion.fecha_finalizacion or curso.fecha_fin

    certificado, _ = Certificado.objects.get_or_create(
        curso=curso,
        alumno=alumno,
    )

    if certificado.estado != Certificado.ESTADO_VALIDO:
        raise PermissionDenied

    codigo = certificado.codigo

    url_verificacion = request.build_absolute_uri(
        reverse(
            "verificar_certificado",
            kwargs={"codigo": codigo},
        )
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="certificado_{codigo}.pdf"'
    )

    ancho, alto = landscape(A4)
    pdf = canvas.Canvas(response, pagesize=(ancho, alto))
    centro = ancho / 2

    azul = colors.HexColor("#2563eb")
    azul_oscuro = colors.HexColor("#172554")
    gris = colors.HexColor("#64748b")
    gris_claro = colors.HexColor("#cbd5e1")
    fondo = colors.HexColor("#f8fafc")

    pdf.setFillColor(fondo)
    pdf.rect(0, 0, ancho, alto, fill=1, stroke=0)

    margen = 16 * mm
    pdf.setFillColor(colors.white)
    pdf.roundRect(
        margen, margen,
        ancho - 2 * margen, alto - 2 * margen,
        5 * mm, fill=1, stroke=0,
    )

    pdf.setStrokeColor(azul)
    pdf.setLineWidth(2.2)
    pdf.roundRect(
        margen, margen,
        ancho - 2 * margen, alto - 2 * margen,
        5 * mm, fill=0, stroke=1,
    )

    pdf.setStrokeColor(colors.HexColor("#dbeafe"))
    pdf.setLineWidth(0.8)
    pdf.roundRect(
        margen + 5 * mm, margen + 5 * mm,
        ancho - 2 * margen - 10 * mm,
        alto - 2 * margen - 10 * mm,
        3 * mm, fill=0, stroke=1,
    )

    # Sello visual de Aula Virtual
    sello_x = centro
    sello_y = alto - 34 * mm
    pdf.setFillColor(colors.HexColor("#eff6ff"))
    pdf.circle(sello_x, sello_y, 8 * mm, fill=1, stroke=0)
    pdf.setFillColor(azul)
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawCentredString(sello_x, sello_y - 1.8 * mm, "AV")

    pdf.setFillColor(azul_oscuro)
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawCentredString(
        centro, alto - 49 * mm, curso.institucion.nombre.upper()
    )

    if curso.institucion.identificacion:
        pdf.setFillColor(gris)
        pdf.setFont("Helvetica", 8.5)
        pdf.drawCentredString(
            centro,
            alto - 55 * mm,
            f"Identificación institucional: {curso.institucion.identificacion}",
        )

    pdf.setFillColor(azul)
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawCentredString(centro, alto - 67 * mm, "AULA VIRTUAL")

    pdf.setFillColor(azul_oscuro)
    pdf.setFont("Helvetica-Bold", 27)
    pdf.drawCentredString(
        centro, alto - 80 * mm, "CERTIFICADO DE APROBACIÓN"
    )

    pdf.setStrokeColor(colors.HexColor("#bfdbfe"))
    pdf.setLineWidth(1)
    pdf.line(centro - 45 * mm, alto - 86 * mm, centro + 45 * mm, alto - 86 * mm)

    pdf.setFillColor(gris)
    pdf.setFont("Helvetica", 12)
    pdf.drawCentredString(centro, alto - 99 * mm, "Se certifica que")

    pdf.setFillColor(azul_oscuro)
    nombre_size = 23 if len(nombre_alumno) <= 38 else 18
    pdf.setFont("Helvetica-Bold", nombre_size)
    pdf.drawCentredString(centro, alto - 114 * mm, nombre_alumno.upper())

    pdf.setFillColor(gris)
    pdf.setFont("Helvetica", 11.5)
    pdf.drawCentredString(
        centro, alto - 127 * mm, "ha aprobado satisfactoriamente el curso"
    )

    pdf.setFillColor(azul)
    curso_size = 20 if len(curso.nombre) <= 45 else 16
    pdf.setFont("Helvetica-Bold", curso_size)
    pdf.drawCentredString(centro, alto - 141 * mm, curso.nombre)

    detalles = []
    if curso.carga_horaria:
        detalles.append(f"Carga horaria: {curso.carga_horaria} horas")
    if fecha_final:
        detalles.append(
            "Finalización: " + fecha_final.strftime("%d/%m/%Y")
        )

    if detalles:
        pdf.setFillColor(gris)
        pdf.setFont("Helvetica", 10)
        pdf.drawCentredString(
            centro, alto - 153 * mm, "   |   ".join(detalles)
        )

    # Firmas: espacios preparados para autoridades institucionales.
    firma_y = 42 * mm
    firma_ancho = 48 * mm
    firma_izq = centro - 70 * mm
    firma_der = centro + 70 * mm

    pdf.setStrokeColor(gris_claro)
    pdf.setLineWidth(0.7)
    pdf.line(
        firma_izq - firma_ancho / 2, firma_y,
        firma_izq + firma_ancho / 2, firma_y,
    )
    pdf.line(
        firma_der - firma_ancho / 2, firma_y,
        firma_der + firma_ancho / 2, firma_y,
    )

    pdf.setFillColor(gris)
    pdf.setFont("Helvetica", 8.5)
    pdf.drawCentredString(
        firma_izq,
        firma_y - 5 * mm,
        "Docente / Responsable",
    )

    autoridad_nombre = curso.institucion.autoridad_nombre.strip()
    autoridad_cargo = curso.institucion.autoridad_cargo.strip()

    if autoridad_nombre:
        pdf.setFillColor(azul_oscuro)
        pdf.setFont("Helvetica-Bold", 8.5)
        pdf.drawCentredString(
            firma_der,
            firma_y - 4.5 * mm,
            autoridad_nombre,
        )

        pdf.setFillColor(gris)
        pdf.setFont("Helvetica", 7.5)
        pdf.drawCentredString(
            firma_der,
            firma_y - 8.5 * mm,
            autoridad_cargo or "Autoridad institucional",
        )
    else:
        pdf.setFillColor(gris)
        pdf.setFont("Helvetica", 8.5)
        pdf.drawCentredString(
            firma_der,
            firma_y - 5 * mm,
            "Autoridad institucional",
        )

    # Pie del certificado
    pdf.setFillColor(colors.HexColor("#94a3b8"))
    pdf.setFont("Helvetica", 7.5)
    pdf.drawCentredString(
        centro,
        27 * mm,
        f"Código de certificado: {codigo}",
    )
    pdf.drawCentredString(
        centro,
        22.5 * mm,
        "Documento generado digitalmente por Aula Virtual",
    )

    # QR de verificación pública, dentro del marco inferior derecho
    qr_widget = qr.QrCodeWidget(url_verificacion)
    qr_bounds = qr_widget.getBounds()
    qr_width = qr_bounds[2] - qr_bounds[0]
    qr_height = qr_bounds[3] - qr_bounds[1]
    qr_size = 18 * mm

    qr_drawing = Drawing(
        qr_size,
        qr_size,
        transform=[
            qr_size / qr_width,
            0,
            0,
            qr_size / qr_height,
            0,
            0,
        ],
    )
    qr_drawing.add(qr_widget)

    qr_x = ancho - margen - qr_size - 8 * mm
    qr_y = margen + 10 * mm

    renderPDF.draw(
        qr_drawing,
        pdf,
        qr_x,
        qr_y,
    )

    pdf.setFillColor(gris)
    pdf.setFont("Helvetica-Bold", 6)
    pdf.drawCentredString(
        qr_x + qr_size / 2,
        qr_y - 3 * mm,
        "VERIFICAR",
    )

    pdf.showPage()
    pdf.save()
    return response


@login_required
def finalizar_curso(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    puede_finalizar = (
        request.user.is_superuser
        or tiene_rol_en_institucion(
            request.user,
            "Coordinador",
            curso.institucion,
        )
        or tiene_rol_en_institucion(
            request.user,
            "Administrador institucional",
            curso.institucion,
        )
    )

    if not puede_finalizar:
        raise PermissionDenied

    inscripciones = Inscripcion.objects.filter(
        curso=curso,
    )

    total_alumnos = inscripciones.count()
    aprobados = inscripciones.filter(
        estado=Inscripcion.ESTADO_APROBADO,
    ).count()
    desaprobados = inscripciones.filter(
        estado=Inscripcion.ESTADO_DESAPROBADO,
    ).count()
    abandonos = inscripciones.filter(
        estado=Inscripcion.ESTADO_ABANDONO,
    ).count()
    inscriptos = inscripciones.filter(
        estado=Inscripcion.ESTADO_INSCRIPTO,
    ).count()
    cursando = inscripciones.filter(
        estado=Inscripcion.ESTADO_CURSANDO,
    ).count()

    pendientes_cierre = inscriptos + cursando
    puede_cerrar = (
        curso.estado != Curso.ESTADO_FINALIZADO
        and pendientes_cierre == 0
    )

    if request.method == "POST":
        if curso.estado == Curso.ESTADO_FINALIZADO:
            messages.info(
                request,
                "El curso ya se encuentra finalizado.",
            )
            return redirect("lista_cursos")

        if pendientes_cierre > 0:
            messages.error(
                request,
                (
                    "No se puede finalizar el curso mientras haya "
                    "alumnos Inscritos o Cursando."
                ),
            )
            return redirect(
                "finalizar_curso",
                pk=curso.pk,
            )

        if request.POST.get("confirmar_finalizacion") != "1":
            messages.error(
                request,
                "Tenés que confirmar la finalización del curso.",
            )
            return redirect(
                "finalizar_curso",
                pk=curso.pk,
            )

        curso.estado = Curso.ESTADO_FINALIZADO
        curso.save(update_fields=["estado"])

        notificar_curso_finalizado(curso)

        messages.success(
            request,
            "Curso finalizado correctamente.",
        )

        return redirect("lista_cursos")

    return render(
        request,
        "cursos/finalizar.html",
        {
            "curso": curso,
            "total_alumnos": total_alumnos,
            "aprobados": aprobados,
            "desaprobados": desaprobados,
            "abandonos": abandonos,
            "inscriptos": inscriptos,
            "cursando": cursando,
            "pendientes_cierre": pendientes_cierre,
            "puede_cerrar": puede_cerrar,
        },
    )


@login_required
def mis_cursos(request):
    inscripciones = (
        request.user.inscripciones
        .select_related(
            "curso",
            "curso__institucion",
        )
        .filter(
            Q(
                estado__in=[
                    Inscripcion.ESTADO_INSCRIPTO,
                    Inscripcion.ESTADO_CURSANDO,
                ]
            )
            | Q(
                curso__estado=Curso.ESTADO_FINALIZADO,
                estado__in=[
                    Inscripcion.ESTADO_APROBADO,
                    Inscripcion.ESTADO_DESAPROBADO,
                    Inscripcion.ESTADO_ABANDONO,
                ],
            )
        )
        .order_by(
            "curso__institucion__nombre",
            "curso__nombre",
        )
    )

    return render(
        request,
        "cursos/mis_cursos.html",
        {
            "inscripciones": inscripciones,
        },
    )


@login_required
def detalle_curso_alumno(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related(
            "institucion"
        ),
        pk=pk,
    )

    estados_permitidos = [
        Inscripcion.ESTADO_INSCRIPTO,
        Inscripcion.ESTADO_CURSANDO,
    ]

    if curso.estado == Curso.ESTADO_FINALIZADO:
        estados_permitidos.extend(
            [
                Inscripcion.ESTADO_APROBADO,
                Inscripcion.ESTADO_DESAPROBADO,
                Inscripcion.ESTADO_ABANDONO,
            ]
        )

    inscripcion = (
        Inscripcion.objects
        .filter(
            curso=curso,
            alumno=request.user,
            estado__in=estados_permitidos,
        )
        .first()
    )

    if not inscripcion:
        raise PermissionDenied

    if not tiene_rol_en_institucion(
        request.user,
        "Alumno",
        curso.institucion,
    ):
        raise PermissionDenied

    ahora = timezone.now()

    clases_visibles = (
        Clase.objects
        .filter(
            visible=True,
        )
        .filter(
            Q(fecha_publicacion__isnull=True)
            | Q(
                fecha_publicacion__lte=ahora
            )
        )
        .order_by(
            "orden",
            "id",
        )
    )

    clases_programadas = (
        Clase.objects
        .filter(
            visible=True,
            fecha_publicacion__gt=ahora,
        )
        .order_by(
            "fecha_publicacion",
            "orden",
            "id",
        )
    )

    modulos = list(
        Modulo.objects
        .filter(
            curso=curso,
            visible=True,
        )
        .prefetch_related(
            Prefetch(
                "clases",
                queryset=clases_visibles,
                to_attr="clases_publicadas",
            ),
            Prefetch(
                "clases",
                queryset=clases_programadas,
                to_attr="clases_programadas",
            ),
        )
        .order_by(
            "orden",
            "id",
        )
    )

    clases_publicadas = []
    proximas_publicaciones = []

    for modulo in modulos:
        clases_publicadas.extend(
            modulo.clases_publicadas
        )

        if modulo.clases_programadas:
            proximas_publicaciones.append(
                modulo.clases_programadas[0].fecha_publicacion
            )

    proxima_publicacion = (
        min(proximas_publicaciones)
        if proximas_publicaciones
        else None
    )

    total_clases = len(
        clases_publicadas
    )

    clases_completadas_ids = set(
        ProgresoClase.objects
        .filter(
            alumno=request.user,
            clase__in=clases_publicadas,
            completada=True,
        )
        .values_list(
            "clase_id",
            flat=True,
        )
    )

    clases_completadas = len(
        clases_completadas_ids
    )

    if total_clases > 0:
        porcentaje_progreso = round(
            (
                clases_completadas
                / total_clases
            )
            * 100
        )
    else:
        porcentaje_progreso = 0

    for modulo in modulos:
        for clase in modulo.clases_publicadas:
            clase.completada_alumno = (
                clase.pk
                in clases_completadas_ids
            )

    return render(
        request,
        "cursos/detalle_alumno.html",
        {
            "curso": curso,
            "inscripcion": inscripcion,
            "modulos": modulos,
            "total_clases": total_clases,
            "clases_completadas": clases_completadas,
            "porcentaje_progreso": porcentaje_progreso,
            "proxima_publicacion": proxima_publicacion,
            "avisos": (
                AvisoCurso.objects
                .filter(
                    curso=curso,
                    visible=True,
                )
                .select_related("autor")[:5]
            ),
        },
    )

@login_required
def seguimiento_curso(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    puede_ver = (
        request.user.is_superuser
        or curso.docentes.filter(pk=request.user.pk).exists()
        or tiene_rol_en_institucion(
            request.user,
            "Coordinador",
            curso.institucion,
        )
    )

    if not puede_ver:
        raise PermissionDenied

    ahora = timezone.now()

    clases_disponibles = list(
        Clase.objects
        .filter(
            modulo__curso=curso,
            modulo__visible=True,
            visible=True,
        )
        .filter(
            Q(fecha_publicacion__isnull=True)
            | Q(fecha_publicacion__lte=ahora)
        )
        .order_by(
            "modulo__orden",
            "orden",
            "id",
        )
    )

    total_clases = len(clases_disponibles)

    actividades_disponibles = list(
        Actividad.objects
        .filter(
            clase__in=clases_disponibles,
            visible=True,
        )
        .select_related("clase")
        .distinct()
    )

    total_actividades = len(actividades_disponibles)

    inscripciones = list(
        Inscripcion.objects
        .filter(curso=curso)
        .select_related("alumno")
        .order_by(
            "alumno__last_name",
            "alumno__first_name",
            "alumno__username",
        )
    )

    alumnos_ids = [
        inscripcion.alumno_id
        for inscripcion in inscripciones
    ]

    progresos = (
        ProgresoClase.objects
        .filter(
            alumno_id__in=alumnos_ids,
            clase__in=clases_disponibles,
            completada=True,
        )
        .values_list("alumno_id", "clase_id")
    )

    clases_por_alumno = {}

    for alumno_id, clase_id in progresos:
        clases_por_alumno.setdefault(
            alumno_id,
            set(),
        ).add(clase_id)

    entregas = (
        Entrega.objects
        .filter(
            alumno_id__in=alumnos_ids,
            actividad__in=actividades_disponibles,
        )
        .select_related("actividad")
    )

    entregas_por_alumno = {}

    for entrega in entregas:
        entregas_por_alumno.setdefault(
            entrega.alumno_id,
            [],
        ).append(entrega)

    seguimiento = []

    for inscripcion in inscripciones:
        alumno = inscripcion.alumno

        completadas = len(
            clases_por_alumno.get(
                alumno.pk,
                set(),
            )
        )

        if total_clases:
            progreso = round(
                completadas
                / total_clases
                * 100
            )
        else:
            progreso = 0

        entregas_alumno = entregas_por_alumno.get(
            alumno.pk,
            [],
        )

        cantidad_entregadas = sum(
            1
            for entrega in entregas_alumno
            if entrega.estado in [
                Entrega.ESTADO_ENTREGADA,
                Entrega.ESTADO_CORREGIDA,
            ]
        )

        cantidad_corregidas = sum(
            1
            for entrega in entregas_alumno
            if entrega.estado == Entrega.ESTADO_CORREGIDA
        )

        porcentajes_calificados = []

        for entrega in entregas_alumno:
            if (
                entrega.calificacion is not None
                and entrega.actividad.puntaje_maximo
                and entrega.actividad.puntaje_maximo > 0
            ):
                porcentaje = (
                    float(entrega.calificacion)
                    / float(
                        entrega.actividad.puntaje_maximo
                    )
                    * 100
                )

                porcentajes_calificados.append(
                    porcentaje
                )

        promedio = None

        if porcentajes_calificados:
            promedio = round(
                sum(porcentajes_calificados)
                / len(porcentajes_calificados)
            )

        seguimiento.append(
            {
                "inscripcion": inscripcion,
                "alumno": alumno,
                "clases_completadas": completadas,
                "progreso": progreso,
                "actividades_entregadas": cantidad_entregadas,
                "actividades_corregidas": cantidad_corregidas,
                "promedio": promedio,
            }
        )

    return render(
        request,
        "cursos/seguimiento.html",
        {
            "curso": curso,
            "seguimiento": seguimiento,
            "total_clases": total_clases,
            "total_actividades": total_actividades,
        },
    )


@login_required
def exportar_seguimiento_excel(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    puede_ver = (
        request.user.is_superuser
        or curso.docentes.filter(pk=request.user.pk).exists()
        or tiene_rol_en_institucion(
            request.user,
            "Coordinador",
            curso.institucion,
        )
    )

    if not puede_ver:
        raise PermissionDenied

    ahora = timezone.now()

    clases_disponibles = list(
        Clase.objects
        .filter(
            modulo__curso=curso,
            modulo__visible=True,
            visible=True,
        )
        .filter(
            Q(fecha_publicacion__isnull=True)
            | Q(fecha_publicacion__lte=ahora)
        )
        .order_by(
            "modulo__orden",
            "orden",
            "id",
        )
    )

    total_clases = len(clases_disponibles)

    actividades_disponibles = list(
        Actividad.objects
        .filter(
            clase__in=clases_disponibles,
            visible=True,
        )
        .select_related("clase")
        .distinct()
    )

    total_actividades = len(actividades_disponibles)

    inscripciones = list(
        Inscripcion.objects
        .filter(curso=curso)
        .select_related("alumno")
        .order_by(
            "alumno__last_name",
            "alumno__first_name",
            "alumno__username",
        )
    )

    alumnos_ids = [
        inscripcion.alumno_id
        for inscripcion in inscripciones
    ]

    progresos = (
        ProgresoClase.objects
        .filter(
            alumno_id__in=alumnos_ids,
            clase__in=clases_disponibles,
            completada=True,
        )
        .values_list("alumno_id", "clase_id")
    )

    clases_por_alumno = {}

    for alumno_id, clase_id in progresos:
        clases_por_alumno.setdefault(
            alumno_id,
            set(),
        ).add(clase_id)

    entregas = (
        Entrega.objects
        .filter(
            alumno_id__in=alumnos_ids,
            actividad__in=actividades_disponibles,
        )
        .select_related("actividad")
    )

    entregas_por_alumno = {}

    for entrega in entregas:
        entregas_por_alumno.setdefault(
            entrega.alumno_id,
            [],
        ).append(entrega)

    filas = []

    for inscripcion in inscripciones:
        alumno = inscripcion.alumno

        completadas = len(
            clases_por_alumno.get(
                alumno.pk,
                set(),
            )
        )

        progreso = (
            round(completadas / total_clases * 100)
            if total_clases
            else 0
        )

        entregas_alumno = entregas_por_alumno.get(
            alumno.pk,
            [],
        )

        cantidad_entregadas = sum(
            1
            for entrega in entregas_alumno
            if entrega.estado in [
                Entrega.ESTADO_ENTREGADA,
                Entrega.ESTADO_CORREGIDA,
            ]
        )

        cantidad_corregidas = sum(
            1
            for entrega in entregas_alumno
            if entrega.estado == Entrega.ESTADO_CORREGIDA
        )

        porcentajes_calificados = []

        for entrega in entregas_alumno:
            if (
                entrega.calificacion is not None
                and entrega.actividad.puntaje_maximo
                and entrega.actividad.puntaje_maximo > 0
            ):
                porcentajes_calificados.append(
                    float(entrega.calificacion)
                    / float(entrega.actividad.puntaje_maximo)
                    * 100
                )

        promedio = (
            round(
                sum(porcentajes_calificados)
                / len(porcentajes_calificados)
            )
            if porcentajes_calificados
            else None
        )

        filas.append(
            [
                alumno.get_full_name() or alumno.username,
                alumno.username,
                inscripcion.get_estado_display(),
                completadas,
                total_clases,
                progreso,
                cantidad_entregadas,
                total_actividades,
                cantidad_corregidas,
                promedio,
            ]
        )

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Seguimiento"

    hoja.append(["SEGUIMIENTO DEL CURSO"])
    hoja.append(["Institución", curso.institucion.nombre])
    hoja.append(["Curso", curso.nombre])
    hoja.append(["Generado", timezone.localtime().strftime("%d/%m/%Y %H:%M")])
    hoja.append([])

    encabezados = [
        "Alumno",
        "Usuario",
        "Estado",
        "Clases completadas",
        "Total clases",
        "Progreso (%)",
        "Actividades entregadas",
        "Total actividades",
        "Actividades corregidas",
        "Promedio (%)",
    ]
    hoja.append(encabezados)

    fila_encabezado = hoja.max_row

    for celda in hoja[fila_encabezado]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    for fila in filas:
        hoja.append(fila)

    hoja.freeze_panes = f"A{fila_encabezado + 1}"
    hoja.auto_filter.ref = (
        f"A{fila_encabezado}:"
        f"J{hoja.max_row}"
    )

    anchos = [32, 20, 18, 20, 14, 15, 24, 18, 24, 16]

    for indice, ancho_columna in enumerate(anchos, start=1):
        hoja.column_dimensions[
            get_column_letter(indice)
        ].width = ancho_columna

    for fila in hoja.iter_rows(
        min_row=fila_encabezado + 1,
        min_col=4,
        max_col=10,
    ):
        for celda in fila:
            celda.alignment = Alignment(horizontal="center")

    nombre_archivo = (
        f"seguimiento_curso_{curso.pk}.xlsx"
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{nombre_archivo}"'
    )

    libro.save(response)
    return response


@login_required
def reporte_cierre_pdf(request, pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=pk,
    )

    puede_ver = (
        request.user.is_superuser
        or curso.docentes.filter(pk=request.user.pk).exists()
        or tiene_rol_en_institucion(
            request.user,
            "Coordinador",
            curso.institucion,
        )
        or tiene_rol_en_institucion(
            request.user,
            "Administrador institucional",
            curso.institucion,
        )
    )

    if not puede_ver:
        raise PermissionDenied

    ahora = timezone.now()

    clases_disponibles = list(
        Clase.objects
        .filter(
            modulo__curso=curso,
            modulo__visible=True,
            visible=True,
        )
        .filter(
            Q(fecha_publicacion__isnull=True)
            | Q(fecha_publicacion__lte=ahora)
        )
        .order_by("modulo__orden", "orden", "id")
    )
    total_clases = len(clases_disponibles)

    actividades_disponibles = list(
        Actividad.objects
        .filter(
            clase__in=clases_disponibles,
            visible=True,
        )
        .select_related("clase")
        .distinct()
    )
    total_actividades = len(actividades_disponibles)

    inscripciones = list(
        Inscripcion.objects
        .filter(curso=curso)
        .select_related("alumno")
        .order_by(
            "alumno__last_name",
            "alumno__first_name",
            "alumno__username",
        )
    )

    alumnos_ids = [
        inscripcion.alumno_id
        for inscripcion in inscripciones
    ]

    progresos = (
        ProgresoClase.objects
        .filter(
            alumno_id__in=alumnos_ids,
            clase__in=clases_disponibles,
            completada=True,
        )
        .values_list("alumno_id", "clase_id")
    )

    clases_por_alumno = {}
    for alumno_id, clase_id in progresos:
        clases_por_alumno.setdefault(alumno_id, set()).add(clase_id)

    entregas = (
        Entrega.objects
        .filter(
            alumno_id__in=alumnos_ids,
            actividad__in=actividades_disponibles,
        )
        .select_related("actividad")
    )

    entregas_por_alumno = {}
    for entrega in entregas:
        entregas_por_alumno.setdefault(
            entrega.alumno_id,
            [],
        ).append(entrega)

    filas = []
    for inscripcion in inscripciones:
        alumno = inscripcion.alumno
        completadas = len(
            clases_por_alumno.get(alumno.pk, set())
        )
        progreso = (
            round(completadas / total_clases * 100)
            if total_clases
            else 0
        )

        entregas_alumno = entregas_por_alumno.get(
            alumno.pk,
            [],
        )
        entregadas = sum(
            1
            for entrega in entregas_alumno
            if entrega.estado in [
                Entrega.ESTADO_ENTREGADA,
                Entrega.ESTADO_CORREGIDA,
            ]
        )

        porcentajes = []
        for entrega in entregas_alumno:
            if (
                entrega.calificacion is not None
                and entrega.actividad.puntaje_maximo
                and entrega.actividad.puntaje_maximo > 0
            ):
                porcentajes.append(
                    float(entrega.calificacion)
                    / float(entrega.actividad.puntaje_maximo)
                    * 100
                )

        promedio = (
            round(sum(porcentajes) / len(porcentajes))
            if porcentajes
            else None
        )

        filas.append([
            alumno.get_full_name() or alumno.username,
            inscripcion.get_estado_display(),
            f"{completadas}/{total_clases}",
            f"{progreso}%",
            f"{entregadas}/{total_actividades}",
            f"{promedio}%" if promedio is not None else "-",
        ])

    total = len(inscripciones)
    aprobados = sum(
        1 for i in inscripciones
        if i.estado == Inscripcion.ESTADO_APROBADO
    )
    desaprobados = sum(
        1 for i in inscripciones
        if i.estado == Inscripcion.ESTADO_DESAPROBADO
    )
    abandonos = sum(
        1 for i in inscripciones
        if i.estado == Inscripcion.ESTADO_ABANDONO
    )
    cursando = sum(
        1 for i in inscripciones
        if i.estado == Inscripcion.ESTADO_CURSANDO
    )
    inscriptos = sum(
        1 for i in inscripciones
        if i.estado == Inscripcion.ESTADO_INSCRIPTO
    )

    response = HttpResponse(
        content_type="application/pdf",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_cierre_curso_{curso.pk}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=32,
        leftMargin=32,
        topMargin=32,
        bottomMargin=32,
        title=f"Reporte académico - {curso.nombre}",
        author="Aula Virtual",
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloReporte",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=rl_colors.HexColor("#172554"),
        spaceAfter=6,
    )
    subtitulo = ParagraphStyle(
        "SubtituloReporte",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        alignment=TA_CENTER,
        textColor=rl_colors.HexColor("#64748b"),
        spaceAfter=14,
    )
    normal = ParagraphStyle(
        "NormalReporte",
        parent=estilos["Normal"],
        fontSize=8,
        leading=11,
        textColor=rl_colors.HexColor("#334155"),
    )

    elementos = [
        Paragraph("REPORTE ACADÉMICO DEL CURSO", titulo),
        Paragraph(
            f"<b>{curso.institucion.nombre}</b><br/>{curso.nombre}",
            subtitulo,
        ),
    ]

    datos_curso = [
        ["Estado", curso.get_estado_display()],
        ["Carga horaria", f"{curso.carga_horaria} h" if curso.carga_horaria else "-"],
        ["Fecha de inicio", curso.fecha_inicio.strftime("%d/%m/%Y") if curso.fecha_inicio else "-"],
        ["Fecha de fin", curso.fecha_fin.strftime("%d/%m/%Y") if curso.fecha_fin else "-"],
        ["Generado", timezone.localtime().strftime("%d/%m/%Y %H:%M")],
    ]
    tabla_datos = Table(datos_curso, colWidths=[105, 370])
    tabla_datos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), rl_colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), rl_colors.HexColor("#172554")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#dbeafe")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.extend([tabla_datos, Spacer(1, 14)])

    resumen = [
        ["Total", "Aprobados", "Desaprobados", "Abandonos", "Cursando", "Inscriptos"],
        [total, aprobados, desaprobados, abandonos, cursando, inscriptos],
    ]
    tabla_resumen = Table(
        resumen,
        colWidths=[79] * 6,
    )
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (-1, 1), rl_colors.HexColor("#172554")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elementos.extend([
        Paragraph("<b>Resumen académico</b>", normal),
        Spacer(1, 6),
        tabla_resumen,
        Spacer(1, 16),
    ])

    encabezados = [
        "Alumno",
        "Estado",
        "Clases",
        "Progreso",
        "Actividades",
        "Promedio",
    ]
    datos_alumnos = [encabezados] + filas

    tabla_alumnos = Table(
        datos_alumnos,
        colWidths=[155, 78, 55, 58, 70, 58],
        repeatRows=1,
    )
    tabla_alumnos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#172554")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            rl_colors.white,
            rl_colors.HexColor("#f8fafc"),
        ]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elementos.extend([
        Paragraph("<b>Detalle de alumnos</b>", normal),
        Spacer(1, 6),
        tabla_alumnos,
    ])

    doc.build(elementos)
    return response


@login_required
def seguimiento_alumno(request, curso_pk, alumno_pk):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_pk,
    )

    puede_gestionar = (
        request.user.is_superuser
        or curso.docentes.filter(pk=request.user.pk).exists()
        or tiene_rol_en_institucion(
            request.user,
            "Coordinador",
            curso.institucion,
        )
        or tiene_rol_en_institucion(
            request.user,
            "Administrador institucional",
            curso.institucion,
        )
    )

    if not puede_gestionar:
        raise PermissionDenied

    inscripcion = get_object_or_404(
        Inscripcion.objects.select_related("alumno"),
        curso=curso,
        alumno_id=alumno_pk,
    )

    estados_gestionables = {
        Inscripcion.ESTADO_CURSANDO,
        Inscripcion.ESTADO_APROBADO,
        Inscripcion.ESTADO_DESAPROBADO,
        Inscripcion.ESTADO_ABANDONO,
    }

    estados_finales = {
        Inscripcion.ESTADO_APROBADO,
        Inscripcion.ESTADO_DESAPROBADO,
        Inscripcion.ESTADO_ABANDONO,
    }

    if (
        request.method == "POST"
        and curso.estado == Curso.ESTADO_FINALIZADO
    ):
        messages.warning(
            request,
            "El curso está finalizado. El estado académico ya no puede modificarse.",
        )
        return redirect(
            "seguimiento_alumno",
            curso_pk=curso.pk,
            alumno_pk=inscripcion.alumno_id,
        )

    if request.method == "POST":
        nuevo_estado = request.POST.get(
            "estado_academico",
            "",
        ).strip()

        if nuevo_estado not in estados_gestionables:
            messages.error(
                request,
                "El estado académico seleccionado no es válido.",
            )

        elif nuevo_estado == inscripcion.estado:
            messages.info(
                request,
                "El alumno ya tiene ese estado académico.",
            )

        elif (
            nuevo_estado in estados_finales
            and request.POST.get("confirmar_estado_final") != "1"
        ):
            messages.error(
                request,
                (
                    "Para asignar un estado final tenés que "
                    "confirmar la decisión."
                ),
            )

        else:
            inscripcion.estado = nuevo_estado

            if nuevo_estado in estados_finales:
                inscripcion.fecha_finalizacion = date.today()
            else:
                inscripcion.fecha_finalizacion = None

            inscripcion.save(
                update_fields=[
                    "estado",
                    "fecha_finalizacion",
                ]
            )

            messages.success(
                request,
                (
                    "Estado académico actualizado a "
                    f"{inscripcion.get_estado_display()}."
                ),
            )

            return redirect(
                "seguimiento_alumno",
                curso_pk=curso.pk,
                alumno_pk=inscripcion.alumno_id,
            )

    alumno = inscripcion.alumno
    ahora = timezone.now()

    clases = list(
        Clase.objects
        .filter(
            modulo__curso=curso,
            modulo__visible=True,
            visible=True,
        )
        .filter(
            Q(fecha_publicacion__isnull=True)
            | Q(fecha_publicacion__lte=ahora)
        )
        .select_related("modulo")
        .order_by("modulo__orden", "orden", "id")
    )

    clases_completadas_ids = set(
        ProgresoClase.objects
        .filter(
            alumno=alumno,
            clase__in=clases,
            completada=True,
        )
        .values_list("clase_id", flat=True)
    )

    detalle_clases = []

    for clase in clases:
        detalle_clases.append(
            {
                "clase": clase,
                "completada": clase.pk in clases_completadas_ids,
            }
        )

    total_clases = len(clases)
    clases_completadas = len(clases_completadas_ids)

    progreso = (
        round(clases_completadas / total_clases * 100)
        if total_clases
        else 0
    )

    actividades = list(
        Actividad.objects
        .filter(
            clase__in=clases,
            visible=True,
        )
        .select_related("clase", "clase__modulo")
        .order_by(
            "clase__modulo__orden",
            "clase__orden",
            "id",
        )
        .distinct()
    )

    entregas = {
        entrega.actividad_id: entrega
        for entrega in Entrega.objects
        .filter(
            alumno=alumno,
            actividad__in=actividades,
        )
        .select_related("actividad")
    }

    detalle_actividades = []
    porcentajes_calificados = []
    actividades_corregidas = 0

    for actividad in actividades:
        entrega = entregas.get(actividad.pk)
        porcentaje_nota = None

        if (
            entrega
            and entrega.estado == Entrega.ESTADO_CORREGIDA
        ):
            actividades_corregidas += 1

        if (
            entrega
            and entrega.calificacion is not None
            and actividad.puntaje_maximo
            and actividad.puntaje_maximo > 0
        ):
            porcentaje_nota = round(
                float(entrega.calificacion)
                / float(actividad.puntaje_maximo)
                * 100
            )

            porcentajes_calificados.append(
                porcentaje_nota
            )

        detalle_actividades.append(
            {
                "actividad": actividad,
                "entrega": entrega,
                "porcentaje_nota": porcentaje_nota,
            }
        )

    total_actividades = len(actividades)

    actividades_entregadas = sum(
        1
        for entrega in entregas.values()
        if entrega.estado in [
            Entrega.ESTADO_ENTREGADA,
            Entrega.ESTADO_CORREGIDA,
        ]
    )

    actividades_rehacer = sum(
        1
        for entrega in entregas.values()
        if entrega.estado == Entrega.ESTADO_REHACER
    )

    actividades_pendientes = max(
        total_actividades - actividades_entregadas,
        0,
    )

    promedio = (
        round(
            sum(porcentajes_calificados)
            / len(porcentajes_calificados)
        )
        if porcentajes_calificados
        else None
    )

    opciones_estado = [
        (
            Inscripcion.ESTADO_CURSANDO,
            "Cursando",
        ),
        (
            Inscripcion.ESTADO_APROBADO,
            "Aprobado",
        ),
        (
            Inscripcion.ESTADO_DESAPROBADO,
            "Desaprobado",
        ),
        (
            Inscripcion.ESTADO_ABANDONO,
            "Abandonó",
        ),
    ]

    return render(
        request,
        "cursos/seguimiento_alumno.html",
        {
            "curso": curso,
            "inscripcion": inscripcion,
            "alumno": alumno,
            "detalle_clases": detalle_clases,
            "total_clases": total_clases,
            "clases_completadas": clases_completadas,
            "progreso": progreso,
            "detalle_actividades": detalle_actividades,
            "total_actividades": total_actividades,
            "actividades_entregadas": actividades_entregadas,
            "actividades_pendientes": actividades_pendientes,
            "actividades_rehacer": actividades_rehacer,
            "actividades_corregidas": actividades_corregidas,
            "promedio": promedio,
            "opciones_estado": opciones_estado,
        },
    )
