from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import (
    get_object_or_404,
    redirect,
    render,
)

from openpyxl import Workbook, load_workbook

from apps.cursos.models import Curso
from apps.roles.utils import tiene_rol_en_institucion
from apps.usuarios.models import (
    MembresiaInstitucional,
    Usuario,
)

from .forms import (
    CargaMasivaAlumnosForm,
    InscripcionForm,
)
from .models import Inscripcion


def puede_gestionar_curso(usuario, curso):
    if usuario.is_superuser:
        return True

    return tiene_rol_en_institucion(
        usuario,
        "Coordinador",
        curso.institucion,
    )


@login_required
def lista_inscripciones(request, curso_id):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_id,
    )

    if not puede_gestionar_curso(
        request.user,
        curso,
    ):
        raise PermissionDenied

    inscripciones = (
        Inscripcion.objects
        .filter(curso=curso)
        .select_related("alumno")
        .order_by(
            "alumno__last_name",
            "alumno__first_name",
            "alumno__username",
        )
    )

    return render(
        request,
        "inscripciones/lista.html",
        {
            "curso": curso,
            "inscripciones": inscripciones,
        },
    )


@login_required
def nueva_inscripcion(request, curso_id):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_id,
    )

    if not puede_gestionar_curso(
        request.user,
        curso,
    ):
        raise PermissionDenied

    if request.method == "POST":
        form = InscripcionForm(
            request.POST,
            curso=curso,
        )

        if form.is_valid():
            inscripcion = form.save(
                commit=False
            )

            inscripcion.curso = curso
            inscripcion.save()

            messages.success(
                request,
                "Alumno inscripto correctamente.",
            )

            return redirect(
                "lista_inscripciones",
                curso_id=curso.pk,
            )

    else:
        form = InscripcionForm(
            curso=curso,
        )

    return render(
        request,
        "inscripciones/form.html",
        {
            "curso": curso,
            "form": form,
            "titulo": "Inscribir alumno",
        },
    )


@login_required
def editar_inscripcion(request, pk):
    inscripcion = get_object_or_404(
        Inscripcion.objects.select_related(
            "curso",
            "curso__institucion",
            "alumno",
        ),
        pk=pk,
    )

    curso = inscripcion.curso

    if not puede_gestionar_curso(
        request.user,
        curso,
    ):
        raise PermissionDenied

    if request.method == "POST":
        form = InscripcionForm(
            request.POST,
            instance=inscripcion,
            curso=curso,
        )

        if form.is_valid():
            form.save()

            messages.success(
                request,
                "Inscripción actualizada correctamente.",
            )

            return redirect(
                "lista_inscripciones",
                curso_id=curso.pk,
            )

    else:
        form = InscripcionForm(
            instance=inscripcion,
            curso=curso,
        )

    return render(
        request,
        "inscripciones/form.html",
        {
            "curso": curso,
            "form": form,
            "titulo": "Editar inscripción",
        },
    )


@login_required
def descargar_plantilla_alumnos(request, curso_id):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_id,
    )

    if not puede_gestionar_curso(
        request.user,
        curso,
    ):
        raise PermissionDenied

    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Alumnos"

    hoja.append(
        [
            "nombre",
            "apellido",
            "usuario",
            "email",
            "contraseña",
        ]
    )

    hoja.append(
        [
            "Juan",
            "Pérez",
            "jperez",
            "juan@ejemplo.com",
            "Temporal123",
        ]
    )

    salida = BytesIO()

    workbook.save(salida)
    salida.seek(0)

    response = HttpResponse(
        salida.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = (
        'attachment; filename="plantilla_alumnos.xlsx"'
    )

    return response


@login_required
def carga_masiva_alumnos(request, curso_id):
    curso = get_object_or_404(
        Curso.objects.select_related("institucion"),
        pk=curso_id,
    )

    if not puede_gestionar_curso(
        request.user,
        curso,
    ):
        raise PermissionDenied

    resultado = None

    if request.method == "POST":
        form = CargaMasivaAlumnosForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():
            archivo = form.cleaned_data["archivo"]

            try:
                workbook = load_workbook(
                    archivo,
                    read_only=True,
                    data_only=True,
                )

                hoja = workbook.active

                filas = list(
                    hoja.iter_rows(
                        values_only=True
                    )
                )

                if not filas:
                    raise ValueError(
                        "El archivo está vacío."
                    )

                encabezados = [
                    str(valor).strip().lower()
                    if valor is not None
                    else ""
                    for valor in filas[0]
                ]

                encabezados_requeridos = [
                    "nombre",
                    "apellido",
                    "usuario",
                    "email",
                    "contraseña",
                ]

                if encabezados != encabezados_requeridos:
                    raise ValueError(
                        "La plantilla no tiene "
                        "las columnas esperadas."
                    )

                creados = 0
                existentes = 0
                inscriptos = 0
                errores = []

                rol_alumno = Group.objects.get(
                    name="Alumno"
                )

                with transaction.atomic():

                    for numero_fila, fila in enumerate(
                        filas[1:],
                        start=2,
                    ):
                        (
                            nombre,
                            apellido,
                            username,
                            email,
                            password,
                        ) = fila

                        nombre = (
                            str(nombre).strip()
                            if nombre
                            else ""
                        )

                        apellido = (
                            str(apellido).strip()
                            if apellido
                            else ""
                        )

                        username = (
                            str(username).strip()
                            if username
                            else ""
                        )

                        email = (
                            str(email).strip()
                            if email
                            else ""
                        )

                        password = (
                            str(password).strip()
                            if password
                            else ""
                        )

                        if not username:
                            errores.append(
                                f"Fila {numero_fila}: "
                                "falta el usuario."
                            )
                            continue

                        usuario = (
                            Usuario.objects
                            .filter(
                                username=username
                            )
                            .first()
                        )

                        if usuario:
                            existentes += 1

                        else:
                            if not password:
                                errores.append(
                                    f"Fila {numero_fila}: "
                                    "falta la contraseña."
                                )
                                continue

                            usuario = Usuario(
                                username=username,
                                first_name=nombre,
                                last_name=apellido,
                                email=email,
                                is_active=True,
                                debe_cambiar_password=True,
                            )

                            usuario.set_password(
                                password
                            )

                            usuario.save()

                            creados += 1

                        membresia, _ = (
                            MembresiaInstitucional.objects
                            .get_or_create(
                                usuario=usuario,
                                institucion=curso.institucion,
                                defaults={
                                    "activa": True,
                                },
                            )
                        )

                        if not membresia.activa:
                            membresia.activa = True

                            membresia.save(
                                update_fields=[
                                    "activa"
                                ]
                            )

                        membresia.roles.add(
                            rol_alumno
                        )

                        _, creada = (
                            Inscripcion.objects
                            .get_or_create(
                                curso=curso,
                                alumno=usuario,
                            )
                        )

                        if creada:
                            inscriptos += 1

                resultado = {
                    "creados": creados,
                    "existentes": existentes,
                    "inscriptos": inscriptos,
                    "errores": errores,
                }

            except Exception as error:
                messages.error(
                    request,
                    (
                        "No se pudo procesar "
                        f"el archivo: {error}"
                    ),
                )

    else:
        form = CargaMasivaAlumnosForm()

    return render(
        request,
        "inscripciones/carga_masiva.html",
        {
            "curso": curso,
            "form": form,
            "resultado": resultado,
        },
    )